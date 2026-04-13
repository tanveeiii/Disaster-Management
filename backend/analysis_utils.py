# analysis_utils.py
from __future__ import annotations

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from flask_cors import CORS
from scipy.stats import linregress, norm
from geopy.distance import geodesic

import geopandas as gpd
from shapely.geometry import LineString
from openpyxl.styles import Border, Side

plt.switch_backend('Agg')
# ---------------------------------------------------------------------
# Basic utilities
# ---------------------------------------------------------------------

from shapely.geometry import Point

def filter_faults_by_radius(gdf_faults: gpd.GeoDataFrame, center_lat: float, center_long: float, radius_km: float) -> gpd.GeoDataFrame:
    """
    Filters a GeoDataFrame of faults to include only those within a specified radius of a location.
    """
    # Create a GeoSeries with the center point
    center_pt = gpd.GeoSeries([Point(center_long, center_lat)], crs="EPSG:4326")
    
    # Project to an Azimuthal Equidistant projection centered on the user's coordinates.
    # This ensures highly accurate distance measurements in meters.
    custom_crs = f"+proj=aeqd +lat_0={center_lat} +lon_0={center_long} +units=m"
    center_proj = center_pt.to_crs(custom_crs)
    faults_proj = gdf_faults.to_crs(custom_crs)
    
    # Calculate the minimum distance from the center point to each fault line
    distances_m = faults_proj.distance(center_proj.iloc[0])
    
    # Keep faults within the specified radius (converting km to meters)
    mask = distances_m <= (radius_km * 1000)
    
    return gdf_faults[mask].copy().reset_index(drop=True)


def filter_earthquakes_by_faults(gdf_eq: gpd.GeoDataFrame, gdf_faults: gpd.GeoDataFrame, buffer_km: float = 15) -> pd.DataFrame:
    """
    Filters earthquake data to include only those falling within a buffer zone of the provided faults.
    """
    if gdf_faults.empty:
        return pd.DataFrame()

    # Project both to EPSG:3857 (Web Mercator) to buffer in meters, matching your existing logic
    gdf_eq_proj = gdf_eq.to_crs(epsg=3857)
    gdf_faults_proj = gdf_faults.to_crs(epsg=3857)
    
    # Create a buffer polygon around each fault line
    buffered_faults = gdf_faults_proj.geometry.buffer(buffer_km * 1000)
    
    # Combine all individual fault buffers into one continuous geographic zone (unary union)
    combined_fault_zones = buffered_faults.unary_union
    
    # Create a mask for earthquakes that fall inside this combined zone
    mask = gdf_eq_proj.within(combined_fault_zones)
    
    # Drop the geometry column and return standard pandas DataFrame for downstream AB/PSHA tasks
    df_filtered = pd.DataFrame(gdf_eq[mask].drop(columns=['geometry']))
    
    return df_filtered.reset_index(drop=True)

def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = df.columns.str.lower().str.strip()
    return df


def convert_to_mw(mag, mag_type):
    if pd.isna(mag):
        return np.nan

    mag_type = str(mag_type).lower()

    if mag_type == "mb":
        return 0.85 * mag + 1.03
    elif mag_type == "ms":
        return 0.67 * mag + 2.07
    elif mag_type == "ml":
        return 0.85 * mag + 0.6
    elif mag_type in ["mw", "m"]:
        return mag
    else:
        return np.nan


def load_sheet(excel_path: str, sheet_name: str) -> pd.DataFrame:
    df = pd.read_excel(excel_path, sheet_name=sheet_name)
    return normalize_columns(df)


# ---------------------------------------------------------------------
# Input + preprocessing
# ---------------------------------------------------------------------

def preprocess_compiled_sheet(excel_path: str, sheet_name: str = "Compiled") -> pd.DataFrame:
    """
    Reads the compiled sheet and creates mw_final.
    """
    df = load_sheet(excel_path, sheet_name)

    if "magnitude" in df.columns and "mag type" in df.columns:
        df["mw_final"] = df.apply(
            lambda r: convert_to_mw(r["magnitude"], r["mag type"]),
            axis=1
        )
    else:
        raise KeyError("Required columns 'magnitude' and 'mag type' not found.")

    return df


def deduplicate_events(df: pd.DataFrame) -> pd.DataFrame:
    """
    Drops duplicates using the same keys as in the notebook.
    Also cleans latitude/longitude strings and converts them to numeric.
    """
    df = df.copy()

    required = ["year", "month", "date", "hours", "minutes", "latitude", "longitude"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise KeyError(f"Missing columns for duplicate removal: {missing}")

    df_nodup = df.drop_duplicates(
        subset=["year", "month", "date", "hours", "minutes", "latitude", "longitude"],
        keep="first"
    ).copy()

    df_nodup["latitude"] = (
        df_nodup["latitude"]
        .astype(str)
        .str.replace(",", "", regex=False)
        .str.strip()
    )
    df_nodup["longitude"] = (
        df_nodup["longitude"]
        .astype(str)
        .str.replace(",", "", regex=False)
        .str.strip()
    )

    df_nodup["latitude"] = pd.to_numeric(df_nodup["latitude"], errors="coerce")
    df_nodup["longitude"] = pd.to_numeric(df_nodup["longitude"], errors="coerce")

    return df_nodup


# ---------------------------------------------------------------------
# Declustering and distance filtering
# ---------------------------------------------------------------------

def decluster(df: pd.DataFrame, dist_km: float = 50) -> pd.DataFrame:
    """
    Keeps the largest-magnitude events while removing nearby events within dist_km.
    """
    df = df.copy().sort_values("mw_final", ascending=False)
    mainshocks = []

    for _, row in df.iterrows():
        keep = True
        for m in mainshocks:
            dist = geodesic(
                (row["latitude"], row["longitude"]),
                (m["latitude"], m["longitude"])
            ).km
            if dist < dist_km:
                keep = False
                break
        if keep:
            mainshocks.append(row)

    return pd.DataFrame(mainshocks).reset_index(drop=True)


def haversine_distance(lat1, lon1, lat2, lon2):
    """
    Vectorized haversine distance in km.
    """
    R = 6371.0

    lat1 = np.radians(lat1)
    lon1 = np.radians(lon1)
    lat2 = np.radians(lat2)
    lon2 = np.radians(lon2)

    dlat = lat2 - lat1
    dlon = lon2 - lon1

    a = np.sin(dlat / 2) ** 2 + np.cos(lat1) * np.cos(lat2) * np.sin(dlon / 2) ** 2
    a = np.clip(a, 0.0, 1.0)
    c = 2 * np.arcsin(np.sqrt(a))

    return R * c


def filter_by_distance(df: pd.DataFrame, center_lat: float, center_long: float, dist_km: float) -> pd.DataFrame:
    distances = haversine_distance(
        center_lat,
        center_long,
        df["latitude"].values,
        df["longitude"].values
    )
    return df[distances <= dist_km].copy().reset_index(drop=True)


# ---------------------------------------------------------------------
# A-B PSHA
# ---------------------------------------------------------------------

def compute_ab_psha(
    df_filtered_final: pd.DataFrame,
    start_year: int = 2020,
    mag_bins=None
):
    """
    Returns:
      yearly_mag_df, cum_df, ab_df, slope, intercept, r2, fig
    """
    if mag_bins is None:
        mag_bins = [
            (3.5, 4.0),
            (4.0, 4.5),
            (4.5, 5.0),
            (5.0, 5.5),
            (5.5, 6.0),
            (6.0, 6.5),
            (6.5, 7.0),
        ]

    end_year = int(df_filtered_final["year"].min())
    years = list(range(start_year, end_year - 1, -1))

    records = []
    for year in years:
        df_y = df_filtered_final[df_filtered_final["year"] == year]
        row = {"Year": year}

        for mmin, mmax in mag_bins:
            label = f"{mmin}-{mmax}"
            count = df_y[
                (df_y["magnitude"] >= mmin) &
                (df_y["magnitude"] < mmax)
            ].shape[0]
            row[label] = count

        records.append(row)

    yearly_mag_df = pd.DataFrame(records)

    cum_df = yearly_mag_df.copy()
    for col in cum_df.columns:
        if col != "Year":
            cum_df[col] = cum_df[col].cumsum()

    last_row = cum_df.iloc[-1]
    data_part = last_row.iloc[1:]
    cum_from_right = data_part[::-1].cumsum()[::-1]

    mag_cols = cum_from_right.index

    M_avg = []
    for col in mag_cols:
        m1, m2 = col.split("-")
        curr_M_avg = (float(m1) + float(m2) - 0.1) / 2
        M_avg.append(curr_M_avg)

    M_avg = pd.Series(M_avg, index=mag_cols, name="M_avg")

    num_years = len(cum_df)
    ab_df = pd.DataFrame({
        "M_avg": M_avg,
        "N_cumulative": cum_from_right,
    })
    ab_df["N"] = ab_df["N_cumulative"] / float(num_years)
    ab_df["logN"] = np.log10(ab_df["N"])
    ab_df = ab_df[ab_df["N"] != 0].copy()

    fig = plt.figure()
    plt.plot(ab_df["M_avg"], ab_df["logN"], marker="o")
    plt.xlabel("Magnitude (M)")
    plt.ylabel("log10(N)")
    plt.title("Gutenberg–Richter Relation")
    plt.grid(True)

    slope, intercept, r_value, p_value, std_err = linregress(
        ab_df["M_avg"], ab_df["logN"]
    )

    b = -slope
    a = intercept
    r2 = r_value ** 2

    return {
        "yearly_mag_df": yearly_mag_df,
        "cum_df": cum_df,
        "ab_df": ab_df,
        "a": a,
        "b": b,
        "r2": r2,
        "figure": fig,
        "slope": slope,
        "intercept": intercept,
        "p_value": p_value,
        "std_err": std_err,
    }


# ---------------------------------------------------------------------
# Completeness analysis
# ---------------------------------------------------------------------

def compute_completeness_analysis(
    df_filtered_final: pd.DataFrame,
    end_year: int = 2020,
    mag_bins=None
) -> pd.DataFrame:
    """
    Returns the same completeness table as in the notebook.
    """
    if mag_bins is None:
        mag_bins = [
            (3.5, 4.0),
            (4.0, 4.5),
            (4.5, 5.0),
            (5.0, 5.5),
            (5.5, 6.0),
            (6.0, 6.5),
            (6.5, 7.0),
        ]

    min_year = int(df_filtered_final["year"].min())
    start_years = list(range(end_year, min_year - 1, -1))

    rows = []
    for start_year in start_years:
        df_period = df_filtered_final[
            (df_filtered_final["year"] >= start_year) &
            (df_filtered_final["year"] <= end_year)
        ]

        Ts = end_year - start_year + 1
        inv_sqrt_Ts = 1 / np.sqrt(Ts)

        row = {
            "Period": start_year,
            "Period Range": f"{start_year}-{end_year}",
            "Ts (Yr)": Ts,
            "(1/Ts)^1/2": round(inv_sqrt_Ts, 3),
        }

        for mmin, mmax in mag_bins:
            label = f"{mmin}-{mmax}" if mmax < 10 else ">6"

            df_m = df_period[
                (df_period["magnitude"] >= mmin) &
                (df_period["magnitude"] < mmax)
            ]

            N = len(df_m)
            lam = N / Ts if Ts > 0 else 0.0
            sigma = np.sqrt(N) / np.sqrt(Ts) if (N > 0 and Ts > 0) else 0.0

            row[f"{label} N"] = N
            row[f"{label} λ"] = round(lam, 5)
            row[f"{label} σs"] = round(sigma, 5)

        rows.append(row)

    return pd.DataFrame(rows)


# ---------------------------------------------------------------------
# Fault processing
# ---------------------------------------------------------------------
import pandas as pd
import geopandas as gpd
from shapely.geometry import LineString

def load_and_build_faults_from_combined(csv_path: str) -> gpd.GeoDataFrame:
    """
    Reads the combined faults CSV, groups the coordinates by S.NO and SEISAT SHEET,
    and builds continuous LineString geometries for each fault.
    """
    # 1. Load the CSV data
    df = pd.read_csv(csv_path, skiprows=1)
    
    # 2. Clean up column names just in case there are trailing spaces
    df.columns = df.columns.str.strip()
    
    lines = []
    metadata = []
    
    # 3. Group by 'SEISAT SHEET' and 'S.NO' 
    # (We group by both because S.NO restarts at 1 for every new sheet)
    grouped = df.groupby(['SEISAT SHEET', 'S.NO'])
    
    for (sheet, sno), group in grouped:
        # Extract the sequence of coordinates for this specific fault
        coords = list(zip(group['LONGITUDE'].astype(float), group['LATITUDE'].astype(float)))
        
        # A LineString requires at least 2 points to be drawn
        if len(coords) >= 2:
            line = LineString(coords)
            lines.append(line)
            
            # Save the metadata from the first row of this group
            first_row = group.iloc[0]
            metadata.append({
                'SEISAT SHEET': sheet,
                'S.NO': sno,
                'Type of Fault': first_row['FAULT TYPE'],
                'NAME OF FAULT (IF ANY)': first_row.get('FAULT NAME', ''),
                'Abbreviation': first_row.get('ABBREVIATION', ''),
                # We save the start and end points in case your downstream formulas still need them
                'LONG 1': coords[0][0],
                'LAT 1': coords[0][1],
                'LONG 2': coords[-1][0],
                'LAT 2': coords[-1][1],
            })
            
    # 4. Create and return the GeoDataFrame
    gdf_faults = gpd.GeoDataFrame(metadata, geometry=lines, crs="EPSG:4326")
    return gdf_faults


def build_geodataframes(df_eq: pd.DataFrame, faults_df: pd.DataFrame):
    """
    Creates GeoDataFrames for earthquakes and faults.
    """
    gdf_eq = gpd.GeoDataFrame(
        df_eq.copy(),
        geometry=gpd.points_from_xy(df_eq.longitude, df_eq.latitude),
        crs="EPSG:4326"
    )

    fault_lines = []
    for _, row in faults_df.iterrows():
        line = LineString([
            (row["LONG 1"], row["LAT 1"]),
            (row["LONG 2"], row["LAT 2"])
        ])
        fault_lines.append(line)

    gdf_faults = gpd.GeoDataFrame(
        faults_df.copy(),
        geometry=fault_lines,
        crs="EPSG:4326"
    )

    return gdf_eq, gdf_faults


def compute_fault_metrics(gdf_eq: gpd.GeoDataFrame, gdf_faults: gpd.GeoDataFrame, buffer_km: float = 15):
    """
    Computes fault lengths, buffered earthquake counts, active length, weights, alpha.
    """
    gdf_eq = gdf_eq.to_crs(epsg=3857)
    gdf_faults = gdf_faults.to_crs(epsg=3857)

    if not gdf_faults.geometry.is_valid.all():
        gdf_faults.loc[~gdf_faults.geometry.is_valid, "geometry"] = (
            gdf_faults.loc[~gdf_faults.geometry.is_valid, "geometry"].buffer(0)
        )

    # Keep the same formula style as your notebook
    R = 6378.8
    lat1 = np.radians(gdf_faults["LAT 1"].astype(float))
    lon1 = np.radians(gdf_faults["LONG 1"].astype(float))
    lat2 = np.radians(gdf_faults["LAT 2"].astype(float))
    lon2 = np.radians(gdf_faults["LONG 2"].astype(float))

    dist_km = R * np.arccos(
        np.sin(lon1) * np.sin(lon2) +
        np.cos(lon1) * np.cos(lon2) * np.cos(lat2 - lat1)
    )
    gdf_faults["Total Length"] = dist_km

    gdf_faults["buffer_geom"] = gdf_faults.geometry.buffer(buffer_km * 1000)

    counts = []
    for _, fault in gdf_faults.iterrows():
        inside = gdf_eq.within(fault["buffer_geom"])
        counts.append(int(inside.sum()))

    gdf_faults["Num_Earthquakes"] = counts

    gdf_faults["Active_Length_km"] = gdf_faults.apply(
        lambda r: r["Total Length"] if r["Num_Earthquakes"] > 0 else 0,
        axis=1
    )
    gdf_faults["Max Mw"] = 5.06 + np.log10(dist_km) * 1.16
    gdf_faults["Wm"] = gdf_faults["Num_Earthquakes"] / 214

    total_active_length = gdf_faults["Active_Length_km"].sum()
    gdf_faults["Wl"] = gdf_faults["Active_Length_km"] / total_active_length if total_active_length != 0 else 0

    gdf_faults["weightage_factor"] = (gdf_faults["Wm"] + gdf_faults["Wl"]) / 2
    gdf_faults["revised_alpha"] = gdf_faults["weightage_factor"] * 2.303

    return gdf_faults


import numpy as np
import pandas as pd
from scipy.stats import norm
from scipy.optimize import curve_fit
from openpyxl.styles import Border, Side
import matplotlib.pyplot as plt

Sln = 0.4648


# ── Geometry helpers ───────────────────────────────────────────────────────────
def get_intersection(x_coord, y_coord, coords):
    best_point = None
    best_dist  = np.inf
    for i in range(len(coords) - 1):
        lat1, lon1 = coords[i]
        lat2, lon2 = coords[i + 1]
        dx = lat2 - lat1
        dy = lon2 - lon1
        if dx == 0 and dy == 0:
            continue
        t  = ((x_coord - lat1) * dx + (y_coord - lon1) * dy) / (dx**2 + dy**2)
        t  = max(0.0, min(1.0, t))
        xi = lat1 + t * dx
        yi = lon1 + t * dy
        dist = np.sqrt((x_coord - xi)**2 + (y_coord - yi)**2)
        if dist < best_dist:
            best_dist  = dist
            best_point = (xi, yi)
    return best_point


def fault_coords_from_row(fault_row):
    geom = fault_row.get("geometry", None)
    if geom is not None and hasattr(geom, "coords"):
        return [(lat, lon) for lon, lat in geom.coords]
    coords = []
    i = 1
    while f"LAT {i}" in fault_row.index and pd.notna(fault_row.get(f"LAT {i}")):
        coords.append((fault_row[f"LAT {i}"], fault_row[f"LONG {i}"]))
        i += 1
    return coords


# ── Core per-fault calculation ─────────────────────────────────────────────────
def calculate_fault(fault_row, x_inter, y_inter, z, x_coord, y_coord, b):
    start_x = fault_row["LAT 1"]
    start_y = fault_row["LONG 1"]
    end_x   = fault_row["LAT 2"]
    end_y   = fault_row["LONG 2"]
    length  = fault_row["Total Length"]
    alpha   = fault_row["revised_alpha"]
    mw_max  = fault_row["Max Mw"]
    m0      = 4
    mag_df  = (mw_max - m0) / 10

    def haversine(lat1, lon1, lat2, lon2):
        cos_term = (
            np.sin(lat1 * np.pi / 180) * np.sin(lat2 * np.pi / 180)
            + np.cos(lat1 * np.pi / 180) * np.cos(lat2 * np.pi / 180)
            * np.cos((lon2 - lon1) * np.pi / 180)
        )
        return 6378.8 * np.arccos(np.clip(cos_term, -1.0, 1.0))

    max_rj = haversine(x_coord, y_coord, start_x, start_y)
    min_rj = haversine(x_coord, y_coord, end_x,   end_y)
    rad_df = (max_rj - min_rj) / 10
    d      = haversine(x_coord, y_coord, x_inter, y_inter)
    l0     = haversine(end_x, end_y, x_inter, y_inter)

    ln_z      = np.log(z)
    rj_values = [min_rj + k * rad_df for k in range(11)]
    table     = []

    for i in range(11):
        m0_curr   = m0 + mag_df * i
        N1 = alpha * (
            (10 ** (-b * (m0_curr - mag_df / 2 - m0)))
            - (10 ** (-b * (mw_max - m0))) / (1 - 10 ** (-b * (mw_max - m0)))
        )
        N2 = alpha * (
            (10 ** (-b * (m0_curr + mag_df / 2 - m0)))
            - (10 ** (-b * (mw_max - m0))) / (1 - 10 ** (-b * (mw_max - m0)))
        )
        lambda_mi = N1 - N2
        Xmi = min(10 ** (-2.44 + 0.59 * m0_curr), length)

        row = {
            ("", "mi"):            round(m0_curr, 3),
            ("", "N(m>mi-Δm/2)"): N1,
            ("", "N(m>mi+Δm/2)"): N2,
            ("", "λ(mi)"):        lambda_mi,
            ("", "Xmi"):          Xmi,
        }
        grouped = {k: {} for k in [
            "P(R<rj+Δr/2 | mi)", "P(R=rj | mi)", "E[ln(Z)]",
            "U", "F'(U)", "P(Z>z | mi)", "µ(z)"
        ]}
        prev_p_less = 0

        for j in rj_values:
            lbl = f"rj = {round(j, 2)}"
            if j < np.sqrt(d**2 + l0**2):
                p_less = 0
            elif j < np.sqrt(d**2 + (length + l0 - Xmi) ** 2):
                p_less = (np.sqrt(j**2 - d**2) - l0) / (length - Xmi)
            else:
                p_less = 1
            grouped["P(R<rj+Δr/2 | mi)"][lbl] = p_less

            p_eq = p_less - prev_p_less
            grouped["P(R=rj | mi)"][lbl] = p_eq
            prev_p_less = p_less

            t1     = -3.7671 + 1.2303 * m0_curr - 0.0019 * m0_curr**2 - 0.0027 * j
            t2     = 1.4857 * np.log(j + 0.0385 * np.exp(0.8975 * m0_curr))
            t3     = 0.1301 * np.log10(j) * max(np.log(j / 100), 0)
            e_ln_z = t1 - t2 + t3 + 0.394
            grouped["E[ln(Z)]"][lbl] = e_ln_z

            u_val     = (ln_z - e_ln_z) / Sln
            grouped["U"][lbl] = u_val

            f_u       = norm.cdf(u_val)
            f_3       = norm.cdf(-3)
            f_prime_u = (f_u - f_3) / (1 - 2 * f_3)
            grouped["F'(U)"][lbl] = f_prime_u

            p_z_gt_z = max(1 - f_prime_u, 0)
            grouped["P(Z>z | mi)"][lbl] = p_z_gt_z
            grouped["µ(z)"][lbl]        = lambda_mi * p_eq * p_z_gt_z

        for metric, vals in grouped.items():
            for lbl, val in vals.items():
                row[(metric, lbl)] = val
        table.append(row)

    df = pd.DataFrame(table)
    df.columns = pd.MultiIndex.from_tuples(df.columns)
    return df, {
        "start_x": start_x, "start_y": start_y,
        "end_x":   end_x,   "end_y":   end_y,
        "x_inter": x_inter, "y_inter": y_inter,
        "length":  length,  "max_rj":  max_rj,
        "min_rj":  min_rj,  "rad_df":  rad_df,
        "d":       d,       "l0":      l0,
        "alpha":   alpha,   "m0":      m0,
        "mw_max":  mw_max,  "mag_df":  mag_df,
    }


# ── Excel helpers ──────────────────────────────────────────────────────────────
def make_header_df(fault_name, p):
    rows = [
        [fault_name, "", "", "", ""],
        ["Starting Point",     "Latitude",  p["start_x"], "Longitude", p["start_y"]],
        ["End Point",          "Latitude",  p["end_x"],   "Longitude", p["end_y"]],
        ["Intersection Point", "Latitude",  p["x_inter"], "Longitude", p["y_inter"]],
        ["", "", "", "", ""],
        ["Length",  p["length"],  "km", "",           ""],
        ["Max rj",  p["max_rj"],  "km", "Radius Dif", p["rad_df"]],
        ["Min rj",  p["min_rj"],  "km", "",           ""],
        ["d",       p["d"],       "km", "",           ""],
        ["Lo",      p["l0"],      "km", "",           ""],
        ["", "", "", "", ""],
        ["Alpha",   p["alpha"],   "",   "",           ""],
        ["Mo",      p["m0"],      "",   "",           ""],
        ["Mw max",  p["mw_max"],  "",   "Mag Dif",    p["mag_df"]],
    ]
    return pd.DataFrame(rows)


def apply_borders(ws, fault_table, current_row):
    right_line  = Side(style="medium")
    bottom_line = Side(style="medium")
    ll0      = fault_table.columns.get_level_values(0)
    end_cols = [i + 2 for i in range(len(ll0) - 1) if ll0[i] != ll0[i + 1]] + [len(ll0) + 1]
    last_row = current_row + fault_table.columns.nlevels + len(fault_table)
    for c in end_cols:
        for r in range(current_row + 1, last_row + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = Border(right=right_line, left=cell.border.left,
                                 top=cell.border.top, bottom=cell.border.bottom)
    for c in range(1, len(ll0) + 2):
        cell = ws.cell(row=last_row, column=c)
        cell.border = Border(bottom=bottom_line, top=cell.border.top,
                             left=cell.border.left, right=cell.border.right)
    return last_row


def power_law(x, a, b_fit):
    return a * np.power(x, b_fit)

#############
def load_faults_from_eq_sheet(excel_path: str, sheet_name: str = "FAULTS") -> gpd.GeoDataFrame:
    faults_df = pd.read_excel(excel_path, sheet_name=sheet_name, header=1)
    faults_df.columns = faults_df.columns.astype(str).str.strip()

    faults_df = faults_df[faults_df["Type of Fault"].notna()].copy()

    coord_cols = ["LONG 1", "LAT 1", "LONG 2", "LAT 2"]
    for col in coord_cols:
        faults_df[col] = pd.to_numeric(faults_df[col], errors="coerce")
    faults_df = faults_df.dropna(subset=coord_cols).copy()

    fault_lines = []
    for _, row in faults_df.iterrows():
        fault_lines.append(
            LineString([
                (row["LONG 1"], row["LAT 1"]),
                (row["LONG 2"], row["LAT 2"])
            ])
        )

    gdf_faults = gpd.GeoDataFrame(
        faults_df,
        geometry=fault_lines,
        crs="EPSG:4326"
    )

    return gdf_faults.reset_index(drop=True)

###########

# ── Flask pipeline ─────────────────────────────────────────────────────────────
def run_full_analysis(
    excel_path: str,
    center_lat: float,
    center_long: float,
    dist_km: float = 300,
    decluster_dist_km: float = 50,
    fault_buffer_km: float = 15,
    start_year: int = 2020,
    x_coord: float = None,
    y_coord: float = None,
    psha_output_path: str = None,   # Excel path for full PSHA export; None = skip writing
):
    # 1. Load and deduplicate
    compiled_df = preprocess_compiled_sheet(excel_path, "Compiled")
    df_nodup    = deduplicate_events(compiled_df)

    gdf_eq_all = gpd.GeoDataFrame(
        df_nodup.copy(),
        geometry=gpd.points_from_xy(df_nodup.longitude, df_nodup.latitude),
        crs="EPSG:4326"
    )

    # 2. Build and filter faults
    gdf_faults_all   = load_and_build_faults_from_combined(
        "India_Faults_Combined.csv"
    )
    gdf_faults_local = filter_faults_by_radius(
        gdf_faults_all, center_lat, center_long, dist_km
    )

    if gdf_faults_local.empty:
        raise ValueError(f"No faults found within {dist_km} km of the specified location.")

    # 3. Filter earthquakes by faults and decluster
    df_filtered_final = filter_earthquakes_by_faults(
        gdf_eq_all, gdf_faults_local, buffer_km=fault_buffer_km
    )
    if df_filtered_final.empty:
        raise ValueError("No earthquakes found within the selected fault buffer.")
    declustered_df = decluster(df_filtered_final, dist_km=decluster_dist_km)

    # 4. G-R analysis and completeness
    ab_results      = compute_ab_psha(df_filtered_final, start_year=start_year)
    completeness_df = compute_completeness_analysis(df_filtered_final, end_year=start_year)

    # 5. Fault metrics
    gdf_eq_local = gpd.GeoDataFrame(
        df_filtered_final.copy(),
        geometry=gpd.points_from_xy(df_filtered_final.longitude, df_filtered_final.latitude),
        crs="EPSG:4326"
    )
    # gdf_faults_metrics = compute_fault_metrics(gdf_eq_local, gdf_faults_local)
    # usage
    gdf_faults_metrics = compute_fault_metrics(
    gdf_eq_local, gdf_faults_local, buffer_km=fault_buffer_km
    )
    outputs = {
        "compiled_df":       compiled_df,
        "df_nodup":          df_nodup,
        "declustered_df":    declustered_df,
        "df_filtered_final": df_filtered_final,
        "local_faults":      gdf_faults_local,
        "gdf_faults":        gdf_faults_metrics,
        "yearly_mag_df":     ab_results["yearly_mag_df"],
        "cum_df":            ab_results["cum_df"],
        "ab_df":             ab_results["ab_df"],
        "a":                 ab_results["a"],
        "b":                 ab_results["b"],
        "r2":                ab_results["r2"],
        "figure":            ab_results["figure"],
        "completeness_df":   completeness_df,
    }

    # 6. Full PSHA — all faults × all z values (0.01 → 0.36)
    if x_coord is None or y_coord is None:
        raise ValueError("x_coord and y_coord are required for PSHA analysis.")

    b        = ab_results["b"]
    z_values = np.round(np.arange(0.01, 0.37, 0.01), 2)
    mu_per_z = []
    writer   = pd.ExcelWriter(psha_output_path, engine="openpyxl") if psha_output_path else None

    for z in z_values:
        sheet_name  = f"{round(z, 2)}g"
        current_row = 0
        total_mu    = 0.0

        for _, fault_row in gdf_faults_metrics.iterrows():
            if fault_row["revised_alpha"] == 0:
                continue
            if (
                pd.isna(fault_row.get("LAT 1")) or
                pd.isna(fault_row.get("LONG 1")) or
                pd.isna(fault_row.get("LAT 2")) or
                pd.isna(fault_row.get("LONG 2"))
            ):
                continue
            coords = [
                (float(fault_row["LAT 1"]), float(fault_row["LONG 1"])),
                (float(fault_row["LAT 2"]), float(fault_row["LONG 2"])),
            ]
            x_inter, y_inter = get_intersection(x_coord, y_coord, coords)

            fault_name  = fault_row.get("NAME OF FAULT (IF ANY)", "Unknown")
            fault_table, params = calculate_fault(
                fault_row, x_inter, y_inter, z, x_coord, y_coord, b
            )

            mu_cols   = [c for c in fault_table.columns if c[0] == "µ(z)"]
            total_mu += fault_table[mu_cols].values.sum()

            if writer is not None:
                header_df = make_header_df(fault_name, params)
                header_df.to_excel(writer, sheet_name=sheet_name,
                                   startrow=current_row, index=False, header=False)
                ws = writer.sheets[sheet_name]
                ws.merge_cells(start_row=current_row + 1, start_column=1,
                               end_row=current_row + 1, end_column=5)
                current_row += len(header_df) + 2
                fault_table.to_excel(writer, sheet_name=sheet_name,
                                     startrow=current_row, merge_cells=True)
                ws.delete_rows(current_row + fault_table.columns.nlevels + 1)
                last_row    = apply_borders(ws, fault_table, current_row)
                current_row = last_row + 5

        mu_per_z.append(total_mu)

    if writer is not None:
        writer.close()
        print(f"Excel written → {psha_output_path}")

    # ── Summary DataFrame ──────────────────────────────────────────────────────
    summary_df = pd.DataFrame({
        "PGA (g)":       z_values,
        "Ln(Z)":         np.log(z_values),
        "Total µ(z)":    mu_per_z,
        "Return Period": 1 / np.array(mu_per_z)
    })
    print(summary_df.to_string(index=False))

    # ── Power Law Fit ──────────────────────────────────────────────────────────
    popt, _ = curve_fit(power_law, z_values, mu_per_z, p0=[1e-6, -2.5], maxfev=10000)
    a_fit, b_fit = popt
    z_fit  = np.linspace(z_values[0], z_values[-1], 300)
    mu_fit = power_law(z_fit, a_fit, b_fit)

    # ── Hazard Curve ───────────────────────────────────────────────────────────
    fig, ax = plt.subplots(figsize=(7, 5))
    ax.plot(summary_df["PGA (g)"], summary_df["Total µ(z)"],
            color="steelblue", linewidth=2, label="Series1")
    ax.plot(z_fit, mu_fit, color="black", linewidth=1,
            label=f"Power (Series1):  y = {a_fit:.0e}x$^{{{b_fit:.3f}}}$")
    ax.set_yscale("log")
    ax.set_xlim(0, 0.4)
    ax.set_xlabel("PGA (g)")
    ax.set_ylabel("Mean Annual Rate of Exceedance")
    ax.set_title("Seismic Hazard Curve for Western\nCentral India (NDMA)")
    ax.legend()
    ax.grid(True, which="both", linestyle="--", alpha=0.5)
    plt.tight_layout()

    # ── MARE Table ─────────────────────────────────────────────────────────────
    scenarios = [
        ("Total µ(z) for 10% probability of exceedance in 50 year",    0.10,  50),
        ("Total µ(z) for 10% probability of exceedance in 100 year",   0.10, 100),
        ("Total µ(z) for 10% probability of exceedance in 500 year",   0.10, 500),
        ("Total µ(z) for 10% probability of exceedance in 1000 year",  0.10, 1000),
        ("Total µ(z) for 2% probability of exceedance in 50 year",     0.02,  50),
        ("Total µ(z) for 2% probability of exceedance in 100 year",    0.02, 100),
        ("Total µ(z) for 2% probability of exceedance in 500 year",    0.02, 500),
        ("Total µ(z) for 2% probability of exceedance in 1000 year",   0.02, 1000),
    ]
    mare_rows = []
    for label, P, t in scenarios:
        lam = -np.log(1 - P) / t
        pga = (lam / a_fit) ** (1 / b_fit)
        mare_rows.append({
            "Scenario":                       label,
            "Mean Annual Rate of Exceedance": lam,
            "PGA (g)":                        round(pga, 3)
        })
    mare_df = pd.DataFrame(mare_rows)

    outputs["psha_summary_df"] = summary_df
    outputs["psha_mare_df"]    = mare_df
    outputs["psha_a_fit"]      = a_fit
    outputs["psha_b_fit"]      = b_fit
    outputs["psha_figure"]     = fig

    return outputs